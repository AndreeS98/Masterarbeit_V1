import streamlit as st
import pandas as pd
from st_aggrid import AgGrid, GridOptionsBuilder, GridUpdateMode
import altair as alt
import pickle
import os
import shutil
import io
from openpyxl import load_workbook

# Funktion zum Laden des Zustands
def load_state():
    if os.path.exists('session_state_shortlist.pkl'):
        with open('session_state_shortlist.pkl', 'rb') as f:
            st.session_state.update(pickle.load(f))

# Funktion zum Speichern des Zustands
def save_state():
    with open('session_state_shortlist.pkl', 'wb') as f:
        pickle.dump(dict(st.session_state), f)

# Zustand laden beim Start
load_state()

# Set initial session state values if they are not already set
if 'intersection_value' not in st.session_state:
    st.session_state['intersection_value'] = 100
if 'stakeholder_importance_value' not in st.session_state:
    st.session_state['stakeholder_importance_value'] = 500
if 'filtered_df' not in st.session_state:
    st.session_state['filtered_df'] = pd.DataFrame()  # Initialize as an empty DataFrame

def Chart(intersection_value, stakeholder_importance_value):
    st.header("Graphische Übersicht")

    if 'selected_columns' in st.session_state and len(st.session_state['selected_columns']) > 0:
        selected_columns = st.session_state['selected_columns']

        # Prepare the data
        if isinstance(selected_columns, list):
            selected_columns_df = pd.DataFrame(selected_columns)
        else:
            selected_columns_df = selected_columns

        columns_to_display = ['Score Finanzen', 'Score Auswirkung']
        selected_columns_df = selected_columns_df[columns_to_display]
        required_columns = ['ID', 'Score Finanzen', 'Score Auswirkung', 'Thema', 'Unterthema', 'Unter-Unterthema', 'Stakeholder Wichtigkeit']

        # Check if the DataFrame is empty after filtering
        if selected_columns_df.empty:
            st.warning("Keine Daten vorhanden, um den Chart anzuzeigen.")
            return  # Stop the function execution

        def assign_color(theme):
            if theme in ['Klimawandel', 'Umweltverschmutzung', 'Wasser- & Meeresressourcen', 'Biodiversität', 'Kreislaufwirtschaft']:
                return 'Environmental'
            elif theme in ['Eigene Belegschaft', 'Belegschaft Lieferkette', 'Betroffene Gemeinschaften', 'Verbraucher und Endnutzer']:
                return 'Social'
            elif theme == 'Unternehmenspolitik':
                return 'Governance'
            else:
                return 'Sonstige'

        selected_columns['color'] = selected_columns['Thema'].apply(assign_color)

        min_rating = st.session_state.combined_df['Stakeholder Gesamtbew.'].min()
        max_rating = st.session_state.combined_df['Stakeholder Gesamtbew.'].max()
        selected_columns['Stakeholder Wichtigkeit'] = ((selected_columns['Stakeholder Gesamtbew.'] - min_rating) / (max_rating - min_rating)) * (1000 - 100) + 100
        selected_columns['Stakeholder Wichtigkeit'] = selected_columns['Stakeholder Wichtigkeit'].fillna(100)

        # Base scatter chart
        scatter = alt.Chart(selected_columns, width=1000, height=800).mark_circle().encode(
            x=alt.X('Score Finanzen', scale=alt.Scale(domain=(0, 1000)), title='Finanzielle Wesentlichkeit'),
            y=alt.Y('Score Auswirkung', scale=alt.Scale(domain=(0, 1000)), title='Auswirkungsbezogene Wesentlichkeit'),
            color=alt.Color('color:N', scale=alt.Scale(
                domain=['Environmental', 'Social', 'Governance', 'Sonstige'],
                range=['green', 'yellow', 'blue', 'gray']
            ), legend=alt.Legend(
                title="Thema",
                orient="right",
                titleColor='black',
                labelColor='black',
                titleFontSize=12,
                labelFontSize=10,
                values=['Environmental', 'Social', 'Governance', 'Sonstige']
            )),
            size=alt.Size('Stakeholder Wichtigkeit:Q', scale=alt.Scale(range=[100, 1000]), legend=alt.Legend(
                title="Stakeholder Wichtigkeit",
                orient="right",
                titleColor='black',
                labelColor='black',
                titleFontSize=12,
                labelFontSize=10
            )),
            tooltip=required_columns
        )

        # Line
        line = alt.Chart(pd.DataFrame({
            'x': [0, st.session_state['intersection_value']],
            'y': [st.session_state['intersection_value'], 0]
        })).mark_line(color='red').encode(
            x='x:Q',
            y='y:Q'
        )

        # Area to the left of the line
        area = alt.Chart(pd.DataFrame({
            'x': [0, 0, st.session_state['intersection_value']],
            'y': [0, st.session_state['intersection_value'], 0]
        })).mark_area(opacity=0.3, color='lightcoral').encode(
            x='x:Q',
            y='y:Q'
        )

        chart = area + scatter + line

        st.altair_chart(chart)
    else:
        st.warning("Keine Daten ausgewählt.")

def filter_table(intersection_value, stakeholder_importance_value):
    st.header("Shortlist")

    if 'selected_columns' in st.session_state and len(st.session_state['selected_columns']) > 0:
        selected_columns = st.session_state['selected_columns']
        
        # Prepare the data
        if isinstance(selected_columns, list):
            selected_columns_df = pd.DataFrame(selected_columns)
        else:
            selected_columns_df = selected_columns
        
        # Filter the data based on the sum of 'Score Finanzen' and 'Score Auswirkung' being greater than intersection_value
        st.session_state.filtered_df = selected_columns_df[
            (selected_columns_df['Score Finanzen'] + selected_columns_df['Score Auswirkung'] > st.session_state['intersection_value']) |
            (selected_columns_df['Stakeholder Wichtigkeit'] > st.session_state['stakeholder_importance_value'])
        ]
        
        # Ensure necessary columns are present
        columns_to_display = ['ID', 'Thema', 'Unterthema', 'Unter-Unterthema', 'Score Finanzen', 'Score Auswirkung']
        filtered_df = st.session_state.filtered_df[columns_to_display]

        if filtered_df.empty:
            st.warning("Keine Inhalte verfügbar")
        else:
            # Configure the grid
            gb = GridOptionsBuilder.from_dataframe(filtered_df)
            gb.configure_side_bar()
            gb.configure_selection('single', use_checkbox=False, groupSelectsChildren="Group checkbox select children", rowMultiSelectWithClick=False)
            grid_options = gb.build()
            
            # Display the grid
            AgGrid(filtered_df, gridOptions=grid_options, enable_enterprise_modules=True, update_mode=GridUpdateMode.MODEL_CHANGED, fit_columns_on_grid_load=True)

    else:
        st.warning("Keine Inhalte vorhanden")

def display_slider():
    st.sidebar.markdown("---")
    # Slider for intersection value
    intersection_value = st.sidebar.slider("Grenzwert für die Relevanz angeben", min_value=0, max_value=1000, value=st.session_state['intersection_value'], step=10)

    # Slider for stakeholder importance value
    stakeholder_importance_value = st.sidebar.slider("Grenzwert für Stakeholder Relevanz angeben", min_value=0, max_value=1000, value=st.session_state['stakeholder_importance_value'], step=50)

    if st.sidebar.button('Auswahl anwenden'):
        st.session_state['intersection_value'] = intersection_value
        st.session_state['stakeholder_importance_value'] = stakeholder_importance_value
        st.session_state['apply_changes'] = True
        st.experimental_rerun()

template_path = os.path.join(os.path.dirname(__file__), 'Templates', 'Ausführung.xlsx')

def transfer_data_to_excel(dataframe):
    # Kopie der Template-Datei erstellen
    temp_excel_path = 'Ausführung.xlsx'
    shutil.copyfile(template_path, temp_excel_path)

    # Laden der Kopie der Excel-Datei
    workbook = load_workbook(temp_excel_path)
    sheet = workbook['Shortlist']

    # Setze die Lasche auf die gewünschte Tabelle
    sheet.title = 'Shortlist'

    first_empty_row = 2

    # Übertragen der Daten in die Excel-Datei
    for index, row in dataframe.iterrows():
        sheet[f'A{first_empty_row}'] = row['Thema']
        sheet[f'B{first_empty_row}'] = row['Unterthema']
        sheet[f'C{first_empty_row}'] = row['Unter-Unterthema']
        first_empty_row += 1

    # Speichern der bearbeiteten Kopie der Excel-Datei
    workbook.save(temp_excel_path)
    st.success('Inhalte erfolgreich zur Excel-Datei hinzugefügt.')

def download_excel():
    # Pfad zur kopierten und bearbeiteten Excel-Datei
    temp_excel_path = 'Ausführung.xlsx'
    workbook = load_workbook(temp_excel_path)
    with io.BytesIO() as virtual_workbook:
        workbook.save(virtual_workbook)
        virtual_workbook.seek(0)
        return virtual_workbook.read()
    
def Excel_button():
    st.sidebar.markdown("---")
    st.sidebar.write("Aktualisieren Sie die Excel-Datei, sofern es Änderungen an der Shortlist gab.")
    if st.sidebar.button('🔃 Excel aktualisieren'):
        transfer_data_to_excel(st.session_state.filtered_df)
    
    # Download-Button für die Excel-Datei
    if st.sidebar.download_button(label="⬇️ Excel-Datei herunterladen",
                          data=download_excel(),
                          file_name="Shortlist.xlsx",
                          mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"):
        st.success("Download gestartet!")

def display_page():
    display_slider()
    if 'apply_changes' in st.session_state and st.session_state['apply_changes']:
        Chart(st.session_state['intersection_value'], st.session_state['stakeholder_importance_value'])
        filter_table(st.session_state['intersection_value'], st.session_state['stakeholder_importance_value'])
    else:
        Chart(100, 500)  # Display initial chart without any filter
    Excel_button()
    save_state()
